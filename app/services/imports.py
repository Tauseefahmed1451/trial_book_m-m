"""Excel input ingestion."""

from __future__ import annotations

import io

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db import models
from app.schemas.books import BookCreate
from app.services.workflow import WorkflowService


class ImportService:
    """Read Excel files and create book workflow records."""

    REQUIRED_COLUMNS = {"title", "notes_on_outline_before"}

    def __init__(self, db: Session) -> None:
        self.db = db

    def import_excel_bytes(self, file_name: str, content: bytes) -> dict[str, object]:
        """Parse uploaded workbook bytes and create books row by row."""
        input_file = models.InputFile(file_name=file_name, source_type="excel", status="uploaded")
        self.db.add(input_file)
        self.db.flush()

        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            input_file.status = "failed"
            return {"imported": 0, "failed": 0, "errors": ["Workbook is empty"]}

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        missing = self.REQUIRED_COLUMNS.difference(headers)
        if missing:
            input_file.status = "failed"
            return {"imported": 0, "failed": 0, "errors": [f"Missing columns: {', '.join(sorted(missing))}"]}

        imported = 0
        failed = 0
        errors: list[str] = []
        workflow = WorkflowService(self.db)

        for row_index, values in enumerate(rows[1:], start=2):
            row_map = {header: values[index] for index, header in enumerate(headers) if header}
            input_row = models.InputRow(
                input_file_id=input_file.id,
                row_number=row_index,
                raw_payload=str(row_map),
                validation_status="valid",
            )
            self.db.add(input_row)

            title = str(row_map.get("title") or "").strip()
            notes = str(row_map.get("notes_on_outline_before") or "").strip()
            if not title or not notes:
                failed += 1
                input_row.validation_status = "invalid"
                input_row.error_message = "title and notes_on_outline_before are required"
                errors.append(f"Row {row_index}: title and notes_on_outline_before are required")
                continue

            payload = BookCreate(
                title=title,
                notes_on_outline_before=notes,
                author_name=self._string_or_none(row_map.get("author_name")),
                audience=self._string_or_none(row_map.get("audience")),
                tone=self._string_or_none(row_map.get("tone")),
                language=self._string_or_none(row_map.get("language")) or "English",
                research_mode=self._string_or_none(row_map.get("research_mode")) or "lightweight",
            )
            book = workflow.create_book(payload, source_type="excel", source_row_ref=str(row_index))
            input_row.validation_status = "imported"
            input_row.book_id = book.id
            imported += 1

        input_file.status = "processed"
        return {"imported": imported, "failed": failed, "errors": errors}

    @staticmethod
    def _string_or_none(value: object) -> str | None:
        text = str(value).strip() if value is not None else ""
        return text or None
